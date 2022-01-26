import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

import fitz
import cv2
from PIL import Image
import os
import pytesseract
import numpy as np

from openpyxl import *
from openpyxl.styles import Alignment
from datetime import datetime


'''
Librerias:
    Openpyxl 3.0.3
    TK-tools 0.12.0
    PyMuPDF 1.17.3
    opencv-python 4.3.0.36
    pytesseract 0.3.4
    Pillow 7.2.0
'''


def is_number(s):
    s = s.replace(' ','')
    try:
        float(s)
        return True
    except ValueError:
        return False

def SegmentoHorizontal(imagen):
    
    global Columna_Excel
    global Fila_Excel
    global wb1
    global ws1
    
    DistanciaFilas = int(entry7.get())
    
    black_line_aux = 0
    
    last_row = 0                    
    
    img_rows_aux = np.sum(imagen,axis=1)
    
    width_crop_aux = imagen.shape[1]
    
    white_rows_aux = 0
    white_row_sum_aux = width_crop_aux*255
                                            
    for i in range(0,len(img_rows_aux)):
        if img_rows_aux[i] == white_row_sum_aux:
            white_rows_aux += 1
        else:
            if white_rows_aux > DistanciaFilas:
                black_line_aux = int(i-(white_rows_aux/2))
                if last_row != 0:
                    img_aux_2 = imagen[last_row:black_line_aux, 0:width_crop_aux]    
                    cv2.imwrite('outfile_aux.png',img_aux_2)
                    text = pytesseract.image_to_string(Image.open('outfile_aux.png'),lang='spa')
                    
                    if len(text) == 0 or text == '':
                        text = pytesseract.image_to_string(Image.open('outfile_aux.png'),lang='spa',config='--psm 10 --oem 3 -c tessedit_char_whitelist=.0123456789')
                                        
#                    print(text)
#                    print(Fila_Excel, Columna_Excel)
                    if is_number(text):
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = float(text.replace(' ',''))
                    else:
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = text
                    
                    if Columna_Excel > 1:
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).alignment = Alignment(horizontal='right')

                    Fila_Excel = Fila_Excel + 1

#                    cv2.imshow('subcrop',img_aux_2)    
#                    cv2.waitKey(0)                
                else:
                    img_aux_2 = imagen[0:black_line_aux,0:width_crop_aux]    
                    cv2.imwrite('outfile_aux.png',img_aux_2)
                    text = pytesseract.image_to_string(Image.open('outfile_aux.png'), lang='spa')
                    
                    if len(text) == 0 or text == '':
                        text = pytesseract.image_to_string(Image.open('outfile_aux.png'),lang='spa',config='--psm 10 --oem 3 -c tessedit_char_whitelist=.0123456789')

#                    print(text)
#                    print(Fila_Excel, Columna_Excel)
                    if is_number(text):
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = float(text.replace(' ',''))
                    else:
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = text
                    
                    if Columna_Excel > 1:
                        ws1.cell(row=Fila_Excel, column=Columna_Excel).alignment = Alignment(horizontal='right')
                    
                    Fila_Excel = Fila_Excel + 1

#                    cv2.imshow('subcrop',img_aux_2)    
#                    cv2.waitKey(0)                

                white_rows_aux = 0
                
                last_row = black_line_aux

    img_aux_2 = imagen[last_row:imagen.shape[0],0:width_crop_aux]    
    cv2.imwrite('outfile_aux.png',img_aux_2)
    text = pytesseract.image_to_string(Image.open('outfile_aux.png'), lang='spa')
    
    if len(text) == 0 or text == '':
        text = pytesseract.image_to_string(Image.open('outfile_aux.png'),lang='spa',config='--psm 10 --oem 3 -c tessedit_char_whitelist=.0123456789')

#    print(text)
#    print(Fila_Excel, Columna_Excel)
    if is_number(text):
        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = float(text.replace(' ',''))
    else:
        ws1.cell(row=Fila_Excel, column=Columna_Excel).value = text
        
    if Columna_Excel > 1:
        ws1.cell(row=Fila_Excel, column=Columna_Excel).alignment = Alignment(horizontal='right')
        
    Fila_Excel = 1

    Columna_Excel = Columna_Excel + 1

#    cv2.imshow('subcrop',img_aux_2)    
#    cv2.waitKey(0)                



#Clase para guardar coordenadas de las rectas, inicio y fin
class Puntos:
    def __init__(self,X1,Y1,X2,Y2):
        self.X1 = X1
        self.Y1 = Y1
        self.X2 = X2
        self.Y2 = Y2



root= tk.Tk()
root.title('ArchivoMex_Extractor')
root.iconbitmap('Settings/ArchivoMex.ico')

canvas1 = tk.Canvas(root, width = 500, height = 380,  relief = 'raised')
canvas1.pack()

label1 = tk.Label(root, text='Extracción de datos')
label1.config(font=('helvetica', 14))
canvas1.create_window(200, 25, window=label1)

label2 = tk.Label(root, text='Nombre del archivo:')
label2.config(font=('helvetica', 10))
canvas1.create_window(75, 80, window=label2)

entry1 = tk.Entry (root) 
entry1.config(state='disabled')
canvas1.create_window(200, 80, window=entry1)


Filename = 'aux'

def seleccionarArchivo():
    global Filename
    root.filename =  filedialog.askopenfilename(initialdir = "",title = "Buscador")
    Filename = str(root.filename)
    entry1.config(state='normal')
    entry1.delete(0,"end")
    entry1.insert(0, Filename)
    entry1.config(state='disabled')
    

button2 = tk.Button(text='Abrir Archivo', command=seleccionarArchivo, bg='white', fg='black', font=('helvetica', 9, 'bold'))
canvas1.create_window(325, 80, window=button2)


label3 = tk.Label(root, text='Número de página:')
label3.config(font=('helvetica', 10))
canvas1.create_window(75, 120, window=label3)

entry2 = tk.Entry (root) 
canvas1.create_window(200, 120, window=entry2)


label4 = tk.Label(root, text='Zoom:')
label4.config(font=('helvetica', 10))
canvas1.create_window(75, 160, window=label4)

entry3 = tk.Entry (root) 
entry3.insert(0, '5')
canvas1.create_window(200, 160, window=entry3)


label4 = tk.Label(root, text='Ancho de Directriz:')
label4.config(font=('helvetica', 10))
canvas1.create_window(75, 200, window=label4)

entry4 = tk.Entry (root) 
entry4.insert(0, '10')
canvas1.create_window(200, 200, window=entry4)

label5 = tk.Label(root, text='Valor recomendado: Zoom * 2')
label5.config(font=('helvetica', 10))
canvas1.create_window(370, 200, window=label5)



label6 = tk.Label(root, text='Distancia columnas')
label6.config(font=('helvetica', 10))
canvas1.create_window(75, 240, window=label6)

entry5 = tk.Entry (root) 
entry5.insert(0, '20')
canvas1.create_window(200, 240, window=entry5)

label7 = tk.Label(root, text='Valor recomendado: Zoom * 4')
label7.config(font=('helvetica', 10))
canvas1.create_window(370, 240, window=label7)


label10 = tk.Label(root, text='Distancia filas')
label10.config(font=('helvetica', 10))
canvas1.create_window(75, 280, window=label10)

entry7 = tk.Entry (root) 
entry7.insert(0, '10')
canvas1.create_window(200, 280, window=entry7)

label11 = tk.Label(root, text='Valor recomendado: Zoom * 2')
label11.config(font=('helvetica', 10))
canvas1.create_window(370, 280, window=label11)


label12 = tk.Label(root, text='Tabla:')
label12.config(font=('helvetica', 10))
canvas1.create_window(300, 120, window=label12)

entry8 = tk.Entry (root) 
entry8.insert(0, '1')
canvas1.create_window(350, 120, window=entry8, width=50)


label12 = tk.Label(root, text='Nombre del Archivo')
label12.config(font=('helvetica', 10))
canvas1.create_window(75, 320, window=label12)

entry9 = tk.Entry (root) 
entry9.insert(0, '')
canvas1.create_window(237, 320, window=entry9, width=200)



wb1 = Workbook()
ws1 = wb1.active
Columna_Excel = 1
Fila_Excel = 1

def extraerDatos():

    if len(str(entry9.get())) < 1:
        messagebox.showinfo(title='Convertidor', message='Introduzca el nombre del archivo')
        return
    
    global Columna_Excel
    global Fila_Excel
    global wb1
    global ws1

    Columna_Excel = 1
    Fila_Excel = 1


    wb1 = Workbook()
    ws1 = wb1.active
    
    Pagina = int(entry2.get())-1
    Zoom = int(entry3.get())
    AnchoDirectriz = int(entry4.get())
    DistanciaColumnas = int(entry5.get())
    ContinuaCoords = 55*Zoom
    Tabla = int(entry8.get())
    
    #abrir PDF en la página indicada
    pdffile = Filename
    doc = fitz.open(pdffile)
    page = doc.loadPage(Pagina) #number of page
    
    #Mejorar la calidad de imagen (Mayor calidad = Mayor tamaño)
    zoom = Zoom    # zoom factor
    mat = fitz.Matrix(zoom, zoom)
    pix = page.getPixmap(matrix = mat)
    
    #Guardar la nueva imagen
    output = "outfile.png"
    pix.writePNG(output)
    
    
    #Abrir OCR
    pytesseract.pytesseract.tesseract_cmd = '/usr/local/Cellar/tesseract/4.1.1/bin/tesseract'
    
    #Abrir imagen y mantener un solo canal, se copia en un nuevo arreglo
    aux = cv2.imread('outfile.png')
    img = aux[:,:,0]
    img_original = img.copy()
    
    #Transformación binaria a la imagen
    img = np.where(img == 0, img, 255)
    
    height = img.shape[0]
    width = img.shape[1]
    
    lengths = [] #Sirve para guardar la longitud de las rectas negras para identificar la mayor
    coordinates = [] #Guarda las coordenadas de inicio y fin de cada recta como un objeto de la clase punto
    
    l = 0 #variable para guardar la longitud de cada recta
    flag = False #indica cuando una recta termina
    last_punto = 0 #indica el último punto (coordenada) para guardar sus coordenadas de fin
    
    #Iterar la imagen en altura y ancho
    for i in range(0,height):
        for j in range(0,width):
            #si la imagen es negra, es una recta
            if img[i][j] == 0:
                #cambiamos la bandera para comenzar a contar, creamos un punto con coordenadas de inicio y se añade al arreglo
                if flag == False:
                    flag = True
                    P  = Puntos(j,i,0,0)
                    coordinates.append(P)
                #sumar la longitud de la recta
                l += 1
            #si la imagen es blanca
            else:
                #si la longitud existe, entonces es el fin de la recta
                if l > 0:
                    #se guarda la longitud de esa recta
                    lengths.append(l)
                    #se guardan las coordenadas de fin del ultimo punto (ultima recta)
                    coordinates[last_punto].X2 = j-1
                    coordinates[last_punto].Y2 = i
                    last_punto += 1
                #reiniciar variables de longitud y bandera
                l = 0
                flag = False
        #si se termina la fila, se reinician variables
        flag = False
        l = 0
    
    #obtener longitud maxima de las rectas
    length_max = np.amax(lengths)
    
#    print(len(lengths))
#    print(len(coordinates))
#    
#    print(length_max)
#    print(coordinates[0].X1,coordinates[0].Y1,coordinates[0].X2,coordinates[0].Y2)
    
    #arreglo y bandera para identificar directrices
    guidelines = []
    flag = False
    
    #arreglo y bandera para identificar encabezados
    guidelines2 = []
    flag2 = False
    
    #se itera el arreglo de longitudes para identificar encabezados y directrices
    for i in range(len(lengths)):
        #si la recta es aproximadamente igual de larga que la recta más larga, entonces es directriz
        if lengths[i]/length_max > .985:
            #si no es la primera recta más larga, se debe de comprobar que no sea un duplicado
            if len(guidelines) > 0:
                #se itera todo el arreglo de directrices para identificar que no sea la misma recta
                for j in range(len(guidelines)):
                    #si la distancia en y es menor a 'n', entonces es la misma recta, cambiamos bandera para identificar que es la misma recta
                    if abs(guidelines[j].Y1-coordinates[i].Y1) < AnchoDirectriz: 
                        flag = True
                #si no es la misma recta, entonces se añade la directriz
                if flag == False:
                    guidelines.append(coordinates[i])
                flag = False
            #como es la primera recta, es directriz
            else:
                guidelines.append(coordinates[i])
        #si la recta no es directriz, pero tiene una longitud considerable, entonces es encabezado
        elif lengths[i] > 100: #***CAMBIAR ESTE VALOR POR UNA VARIABLE
            #buscar duplicados, comparar alturas
            if len(guidelines2) > 0:
                for j in range(len(guidelines2)):
                    if abs(guidelines2[j].X1-coordinates[i].X1) < AnchoDirectriz: 
                        flag2 = True
                if flag2 == False:
                    guidelines2.append(coordinates[i])
                flag2 = False
            #si es el primer encabezado se añade
            else:
                guidelines2.append(coordinates[i])
                
        flag2 = False
        flag = False
    
    
    #Si en el PDF la hoja dice continua, se añade una directriz "invisible" para detener la segmentación
    if len(guidelines) < 3:
        P  = Puntos(guidelines[0].X1,height-ContinuaCoords,guidelines[0].X2,height-ContinuaCoords)
        guidelines.append(P)
    
#    print('Lineas principales')
#    for i in range(len(guidelines)):
#        print(guidelines[i].X1,guidelines[i].Y1,guidelines[i].X2,guidelines[i].Y2)
#    
#    print('Lineas encabezado')
#    for i in range(len(guidelines2)):
#        print(guidelines2[i].X1,guidelines2[i].Y1,guidelines2[i].X2,guidelines2[i].Y2)
    
    cv2.imwrite('outfile2.png',img)


    #---------------------------------------------------
    
    #recortar la imagen a partir de la segunda y última directriz, estos son los datos
    indiceTabla = (Tabla-1)*3
    crop_img = img_original[guidelines[1+indiceTabla].Y1:guidelines[2+indiceTabla].Y2, guidelines[1+indiceTabla].X1:guidelines[1+indiceTabla].X2].copy()
#    cv2.imshow("cropped", crop_img)
#    cv2.waitKey(0)
    cv2.imwrite('cropimg.png',crop_img)
    
    #Suma los valores totales de cada fila
    img_rows = np.sum(crop_img,axis=1)
    
    height_crop = crop_img.shape[0]
    width_crop = crop_img.shape[1]
    
    white_rows = 0
    white_row_sum = width_crop*255
    white_column_sum = height_crop*255
    
    
    #itera todas las filas
    for i in range(0,len(img_rows)):
        #si la suma total es igual al color de pixel por ancho, entonces es una recta
        if img_rows[i] == crop_img[i][0]*width_crop:
            #se pinta la recta de blanco
            crop_img[i,:] = 255
    
    #reiniciar sumas totales
    img_rows = np.sum(crop_img,axis=1)
    img_columns = np.sum(crop_img,axis=0)
    
    cv2.imwrite("cropimg_Nonoise.png", crop_img)
    
    #Metodo para obtener la información por filas
    #Segmenta la imagen horizontalmente entre cada renglon de la tabla
    #Puede presentar errores 
    '''
    for i in range(0,len(img_rows)):
        if img_rows[i] == white_row_sum:
            white_rows += 1
        else:
            if white_rows > 3:
                black_line = int(i-(white_rows/2))
                crop_img[black_line,:] = 0
            white_rows = 0
    '''

    #---------------------------------------------------
    
    #Arreglo para guardar los puntos entre columnas de datos
    Data = []
    
    last_column = 0
    
    #iterar todas las columnas de la imagen
    for i in range(0,len(img_columns)):
        #si la columna es blanca, sumamos al contador
        if img_columns[i] == white_column_sum:
            white_rows += 1
        #si la columna no es completamente blanca
        else:
            #si existe una 'distancia' considerable, significa que acabamos una columna
            if white_rows > DistanciaColumnas: 
                #Se marca el corte entre un punto medio entre columnas de datos
                black_line = int(i-(white_rows/2))
                #si la última columna es distinta de cero, es decir, no es la primera
                if last_column != 0:
                    #creamos un punto con las coordenadas de ese corte
                    D = Puntos(last_column,0,black_line,height_crop)
                    Data.append(D)
                    #recortamos la imagen con la columna de datos y se extrae el texto
                    img_aux = crop_img[0:height_crop, last_column:black_line]
                    
                    SegmentoHorizontal(img_aux)
                                        
                    cv2.imwrite('outfile3.png',img_aux)
                    #text = pytesseract.image_to_string(Image.open('outfile3.png'))
#                    print(text)
#                    cv2.imshow("cropped", img_aux)
#                    cv2.waitKey(0)                
                #si es la primera columna
                else:
                    #creamos un punto con las coordenadas
                    D = Puntos(0,0,black_line,height_crop)
                    Data.append(D)
                    #extraemos el texto
                    img_aux = crop_img[0:height_crop, 0:black_line]
                    
                    SegmentoHorizontal(img_aux)

                    cv2.imwrite('outfile3.png',img_aux)
                    #text = pytesseract.image_to_string(Image.open('outfile3.png'))
#                    print(text)
#                    cv2.imshow("cropped", img_aux)
#                    cv2.waitKey(0)
                
                #se actualiza la última columna en las coordenadas de corte
                last_column = black_line
                
                #black_line = int(i-(white_rows/2))
                #crop_img[:,black_line] = 0
            white_rows = 0
    
    #como la última columna está pegada al final de la imagen, la extracción se hace manual
    #y se añade su coordenada de corte 
    img_aux = crop_img[0:height_crop, last_column:width_crop]
    
    SegmentoHorizontal(img_aux)
    
    cv2.imwrite('outfile3.png',img_aux)
    #text = pytesseract.image_to_string(Image.open('outfile3.png'))
#    print(text)
#    cv2.imshow("cropped", img_aux)
#    cv2.waitKey(0)                
    Last_Column_Data = Puntos(last_column,0,width_crop,height_crop)
    Data.append(Last_Column_Data)


    #---------------------------------------------------

    nombre = str(entry9.get()).replace('<','').replace('>','').replace(':','').replace('"','').replace('/','').replace('\\','').replace('|','').replace('?','').replace('*','').replace('\n',' ')
    fecha = datetime.now().strftime('%Y-%m-%d %H-%M-%S')

    if len(nombre) < 1:
        nombre = 'ArchivoMex ' + fecha 
        
    wb1.save(nombre + ".xlsx")
    wb1.close()

    Columna_Excel = 1
    Fila_Excel = 1

    messagebox.showinfo(title='Convertidor', message='Completado!')

button3 = tk.Button(text='Procesar', command=extraerDatos, bg='white', fg='black', font=('helvetica', 9, 'bold'))
canvas1.create_window(200, 360, window=button3)

root.lift()
root.mainloop()